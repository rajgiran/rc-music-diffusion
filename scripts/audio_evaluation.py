#!/usr/bin/env python3
"""
Audio Model Evaluation Suite
Compares AudioLDM1 vs AudioLDM2 using objective & subjective metrics,
adds FAD (VGGish) and CLAP prompt adherence (Transformers).

Usage example:
  python audio_evaluation.py \
    --model1-dir /path/to/audioldm1_wavs \
    --model2-dir /path/to/audioldm2_wavs \
    --prompts-file /path/to/prompts.txt \
    --output-dir /path/to/eval_out \
    --model1-name AudioLDM1 --model2-name AudioLDM2 \
    --fad-device cpu \
    --clap-checkpoint laion/clap-htsat-unfused \
    --clap-device cuda
"""

import os
import json
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
import librosa
import soundfile as sf
import matplotlib.pyplot as plt
import seaborn as sns

from scipy import stats
from scipy.spatial.distance import pdist, squareform
from scipy.linalg import sqrtm

warnings.filterwarnings("ignore")

# Optional TensorFlow (for FAD / VGGish)
try:
    import tensorflow as tf
    import tensorflow_hub as hub
    TF_AVAILABLE = True
except Exception:
    TF_AVAILABLE = False

# CLAP via Transformers (robust; no custom CUDA ops)
try:
    import torch
    from transformers import AutoProcessor, ClapModel
    HF_CLAP_AVAILABLE = True
except Exception:
    HF_CLAP_AVAILABLE = False


# ----------------------------
# Helpers
# ----------------------------
def load_audio_mono(path, target_sr=None):
    y, sr = librosa.load(path, sr=target_sr, mono=True)
    if not np.all(np.isfinite(y)):
        y = np.nan_to_num(y)
    return y, sr


def rms_db(x, eps=1e-12):
    r = np.sqrt(np.mean(x * x) + eps)
    return 20.0 * np.log10(r + eps)


def safe_div(a, b, eps=1e-12):
    return a / (b + eps)


def quick_snr_estimate(y, frame_length=2048, hop_length=512, eps=1e-12):
    """Heuristic SNR: estimate noise as the 10th percentile of frame RMS."""
    frame_rms = librosa.feature.rms(y=y, frame_length=frame_length, hop_length=hop_length).squeeze(0)
    if frame_rms.size == 0:
        return 0.0
    signal_rms = np.mean(frame_rms)
    noise_rms = np.percentile(frame_rms, 10)
    snr = 20.0 * np.log10(safe_div(signal_rms, noise_rms, eps))
    return float(snr)


def compute_objective_features(path, sr_analysis=22050):
    """Compute per-file objective features used in industry-typical quick screens."""
    try:
        y, sr = load_audio_mono(path, target_sr=sr_analysis)
        duration = len(y) / float(sr)

        # RMS (linear) and dB
        rms_lin = float(np.sqrt(np.mean(y ** 2)))
        rms_db_val = float(rms_db(y))

        # Peak
        peak = float(np.max(np.abs(y)) + 1e-12)

        # SNR (heuristic)
        snr = quick_snr_estimate(y)

        # Spectral features
        S = np.abs(librosa.stft(y, n_fft=2048, hop_length=512)) ** 2
        spec_centroid = float(np.mean(librosa.feature.spectral_centroid(S=S, sr=sr)))
        spec_bandwidth = float(np.mean(librosa.feature.spectral_bandwidth(S=S, sr=sr)))
        # Flatness ~ noise-like vs tone-like; reused later as "frequency_diversity"
        spec_flatness = float(np.mean(librosa.feature.spectral_flatness(S=S)))

        # Dynamic range in dB from frame RMS percentiles
        frame_rms = librosa.feature.rms(S=S).squeeze(0)
        if frame_rms.size > 0:
            p95 = np.percentile(frame_rms, 95)
            p05 = np.percentile(frame_rms, 5)
            dyn_range = float(20.0 * np.log10(safe_div(p95, p05)))
        else:
            dyn_range = 0.0

        return dict(
            file=str(path),
            duration=duration,
            rms=rms_lin,
            rms_db=rms_db_val,
            peak=peak,
            snr=snr,
            spectral_centroid=spec_centroid,
            spectral_bandwidth=spec_bandwidth,
            frequency_diversity=spec_flatness,
            dynamic_range=dyn_range,
        )
    except Exception as e:
        return dict(file=str(path), error=str(e))


# ----------------------------
# Evaluator
# ----------------------------
class AudioModelEvaluator:
    """Industry-standard evaluation suite with FAD and human evaluation protocols"""

    def __init__(
        self,
        model1_dir,
        model2_dir,
        prompts_file,
        model1_name="AudioLDM1",
        model2_name="AudioLDM2",
        fad_device="cpu",
        clap_checkpoint="laion/clap-htsat-unfused",
        clap_device="cuda",
    ):
        self.model1_dir = Path(model1_dir)
        self.model2_dir = Path(model2_dir)
        self.model1_name = model1_name
        self.model2_name = model2_name

        # Settings
        self.fad_device = fad_device  # "cpu" (default) or "gpu"
        self.fad_target_sr = 16000
        self.clap_checkpoint = clap_checkpoint
        self.clap_device = clap_device

        # Load prompts for semantic evaluation
        with open(prompts_file, "r") as f:
            self.prompts = [line.strip() for line in f if line.strip()]

        # Get audio files
        self.model1_files = sorted(list(self.model1_dir.glob("*.wav")))
        self.model2_files = sorted(list(self.model2_dir.glob("*.wav")))

        print(f"[{self.model1_name}] {len(self.model1_files)} files")
        print(f"[{self.model2_name}] {len(self.model2_files)} files")

        # Models (lazy-loaded)
        self.vggish_model = None
        self.clap_model = None
        self.clap_processor = None

        # Load VGGish model for FAD calculation
        if TF_AVAILABLE:
            self._load_vggish_model()
        else:
            print("[FAD] Disabled (TensorFlow not available)")

        # Load CLAP model (Transformers)
        if HF_CLAP_AVAILABLE:
            self._load_clap_model()
        else:
            print("[CLAP] Disabled (Transformers CLAP not available)")

    # ----------------------------
    # FAD (VGGish via TF-Hub)
    # ----------------------------
    def _load_vggish_model(self):
        """Load TF-Hub VGGish; optionally force CPU to dodge CUDA/cuDNN conflicts."""
        try:
            if self.fad_device == "cpu":
                try:
                    tf.config.set_visible_devices([], "GPU")
                except Exception:
                    pass
                os.environ["CUDA_VISIBLE_DEVICES"] = "-1"

            print("[FAD] Loading VGGish from TF-Hub …")
            self.vggish_model = hub.load("https://tfhub.dev/google/vggish/1")
            print("[FAD] VGGish ready")
        except Exception as e:
            print(f"[FAD] Disabled (load error: {e})")
            self.vggish_model = None

    def extract_vggish_embeddings(self, audio_files):
        """Extract VGGish embeddings for FAD (expects 1-D waveform)."""
        if self.vggish_model is None:
            return None

        embs = []
        for p in audio_files:
            try:
                y, sr = load_audio_mono(p, target_sr=self.fad_target_sr)

                # VGGish needs at least 0.96s
                min_len = int(0.96 * self.fad_target_sr)
                if len(y) < min_len:
                    y = np.pad(y, (0, min_len - len(y)))

                # 1-D float32 tensor (no batch dim!)
                x = tf.convert_to_tensor(y, dtype=tf.float32)
                e = self.vggish_model(x).numpy()  # shape (T, 128) or similar
                if e.ndim == 2:
                    e = e.mean(axis=0)
                embs.append(e[None, :])
            except Exception as e:
                print(f"[FAD] embed fail {Path(p).name}: {e}")
        return np.vstack(embs) if embs else None

    def calculate_fad(self, embeddings1, embeddings2):
        """Calculate Fréchet Audio Distance between two sets of embeddings"""
        if embeddings1 is None or embeddings2 is None:
            return None

        mu1 = np.mean(embeddings1, axis=0)
        mu2 = np.mean(embeddings2, axis=0)
        sigma1 = np.cov(embeddings1, rowvar=False)
        sigma2 = np.cov(embeddings2, rowvar=False)

        diff = mu1 - mu2
        eps = 1e-6
        sigma1 = sigma1 + eps * np.eye(sigma1.shape[0])
        sigma2 = sigma2 + eps * np.eye(sigma2.shape[0])

        try:
            sqrt_product = sqrtm(sigma1.dot(sigma2))
            if np.iscomplexobj(sqrt_product):
                sqrt_product = sqrt_product.real
        except Exception:
            sqrt_product = np.eye(sigma1.shape[0])

        fad = np.sum(diff ** 2) + np.trace(sigma1 + sigma2 - 2 * sqrt_product)
        return float(fad)

    def evaluate_fad_score(self):
        """Evaluate FAD between two models"""
        if self.vggish_model is None:
            print("[FAD] Skipped (no model).")
            return None

        print("[Eval] FAD …")
        emb1 = self.extract_vggish_embeddings(self.model1_files)
        emb2 = self.extract_vggish_embeddings(self.model2_files)
        if emb1 is None or emb2 is None:
            print("[FAD] Skipped (no embeddings).")
            return None

        fad_score = self.calculate_fad(emb1, emb2)
        return {
            "fad_score": fad_score,
            "model1_samples": int(emb1.shape[0]),
            "model2_samples": int(emb2.shape[0]),
            "interpretation": "Lower FAD indicates more similar distributions",
        }

    # ----------------------------
    # CLAP (Transformers)
    # ----------------------------
    def _load_clap_model(self):
        """Load CLAP via Transformers (supports laion/clap-htsat-unfused/fused)."""
        try:
            print("[CLAP] Loading via Transformers …")
            self.clap_processor = AutoProcessor.from_pretrained(self.clap_checkpoint)
            self.clap_model = ClapModel.from_pretrained(self.clap_checkpoint)
            dev = self.clap_device
            if dev == "cuda" and not torch.cuda.is_available():
                dev = "cpu"
            self.clap_device = dev
            self.clap_model.to(dev)
            self.clap_model.eval()
            print(f"[CLAP] Loaded: {self.clap_checkpoint} on {dev}")
        except Exception as e:
            self.clap_model = None
            self.clap_processor = None
            print(f"[CLAP] Disabled (load error: {e})")

    @torch.inference_mode()
    def clap_text_audio_similarity(self, wav_path: Path, prompt: str, sr_target=48000):
        if self.clap_model is None:
            return None
        y, _ = librosa.load(wav_path, sr=sr_target, mono=True)
        inputs = self.clap_processor(
            text=[prompt],
            audios=[y],
            return_tensors="pt",
            sampling_rate=sr_target,
            padding=True,
        )
        inputs = {k: v.to(self.clap_device) for k, v in inputs.items()}
        outs = self.clap_model(**inputs)
        # Normalize and cosine sim
        at = outs.audio_embeds / outs.audio_embeds.norm(dim=-1, keepdim=True)
        tt = outs.text_embeds / outs.text_embeds.norm(dim=-1, keepdim=True)
        sim = (at @ tt.T).squeeze().item()
        return float(sim)

    def evaluate_prompt_adherence(self, output_dir: Path):
        """Compute CLAP text↔audio cosine similarity per file and save CSV/JSON."""
        if self.clap_model is None:
            print("[CLAP] Skipped (no model).")
            return None

        print("[Eval] CLAP prompt adherence …")
        rows = []
        # Pair files and prompts by index
        n1 = len(self.model1_files)
        n2 = len(self.model2_files)
        n = min(n1, n2, len(self.prompts))
        for i in range(n):
            p = self.prompts[i]
            f1 = self.model1_files[i]
            f2 = self.model2_files[i]
            try:
                s1 = self.clap_text_audio_similarity(f1, p)
            except Exception as e:
                s1 = None
                print(f"[CLAP] fail {f1.name}: {e}")
            try:
                s2 = self.clap_text_audio_similarity(f2, p)
            except Exception as e:
                s2 = None
                print(f"[CLAP] fail {f2.name}: {e}")

            rows.append(dict(index=i, model=self.model1_name, file=f1.name, prompt=p, clap_similarity=s1))
            rows.append(dict(index=i, model=self.model2_name, file=f2.name, prompt=p, clap_similarity=s2))

        df = pd.DataFrame(rows)
        df.to_csv(output_dir / "clap_prompt_adherence.csv", index=False)
        with open(output_dir / "clap_prompt_adherence.json", "w") as f:
            json.dump(rows, f, indent=2)
        return df

    # ----------------------------
    # Objective quality + diversity
    # ----------------------------
    def evaluate_audio_quality(self, audio_files, model_name):
        """Compute objective metrics for a list of audio files."""
        print("[Eval] Objective features …")
        feats = []
        for f in audio_files:
            m = compute_objective_features(f, sr_analysis=22050)
            m["model"] = model_name
            feats.append(m)
        df = pd.DataFrame(feats)
        return df

    def evaluate_audio_diversity(self, audio_files, model_name):
        """Measure diversity across generated samples using mel-spec stats."""
        print("[Eval] Diversity …")
        embeddings = []
        for file in audio_files:
            try:
                audio, sr = librosa.load(file, sr=16000, mono=True)
                mel = librosa.feature.melspectrogram(y=audio, sr=sr, n_mels=128)
                mel_db = librosa.power_to_db(mel, ref=np.max)

                # Mean, std, max over time for each mel bin
                features = np.concatenate(
                    [mel_db.mean(axis=1), mel_db.std(axis=1), mel_db.max(axis=1)]
                )
                embeddings.append(features)
            except Exception as e:
                print(f"[Diversity] fail {Path(file).name}: {e}")
                continue

        if len(embeddings) < 2:
            return {
                "model": model_name,
                "mean_pairwise_distance": np.nan,
                "std_pairwise_distance": np.nan,
                "min_distance": np.nan,
                "max_distance": np.nan,
                "diversity_score": np.nan,
                "num_samples": len(embeddings),
            }

        embeddings = np.array(embeddings)
        distances = pdist(embeddings, metric="cosine")
        return {
            "model": model_name,
            "mean_pairwise_distance": float(np.mean(distances)),
            "std_pairwise_distance": float(np.std(distances)),
            "min_distance": float(np.min(distances)),
            "max_distance": float(np.max(distances)),
            "diversity_score": float(np.mean(distances)),  # Higher = more diverse
            "num_samples": int(len(embeddings)),
        }

    # ----------------------------
    # Statistical comparison
    # ----------------------------
    def compare_models_statistically(self, df1, df2, metrics):
        """Statistical comparison between models."""
        print("[Eval] Statistical tests …")
        comparison_results = {}
        for metric in metrics:
            if metric in df1.columns and metric in df2.columns:
                values1 = df1[metric].dropna()
                values2 = df2[metric].dropna()
                if len(values1) == 0 or len(values2) == 0:
                    continue

                statistic, p_value = stats.ttest_ind(values1, values2, equal_var=False)
                pooled_std = np.sqrt(
                    ((len(values1) - 1) * values1.var() + (len(values2) - 1) * values2.var())
                    / max(len(values1) + len(values2) - 2, 1)
                )
                cohens_d = (values1.mean() - values2.mean()) / pooled_std if pooled_std > 0 else 0.0

                comparison_results[metric] = {
                    f"{self.model1_name}_mean": float(values1.mean()),
                    f"{self.model1_name}_std": float(values1.std()),
                    f"{self.model2_name}_mean": float(values2.mean()),
                    f"{self.model2_name}_std": float(values2.std()),
                    "p_value": float(p_value),
                    "cohens_d": float(cohens_d),
                    "significant": bool(p_value < 0.05),
                    "effect_size": "small" if abs(cohens_d) < 0.5 else ("medium" if abs(cohens_d) < 0.8 else "large"),
                }
        return comparison_results

    # ----------------------------
    # Human evaluation assets
    # ----------------------------
    def create_human_evaluation_protocol(self, output_dir):
        """Create comprehensive human listening test protocol and materials."""
        print("[Eval] Human evaluation materials …")
        test_pairs = self._create_ab_test_pairs()
        self._create_evaluation_forms(output_dir, test_pairs)
        self._create_analysis_templates(output_dir)
        self._create_experimenter_guidelines(output_dir)
        return test_pairs

    def _create_ab_test_pairs(self):
        """Create randomized A/B test pairs for human evaluation."""
        import random

        test_pairs = []
        min_files = min(len(self.model1_files), len(self.model2_files), len(self.prompts))
        for i in range(min_files):
            for criterion in ["overall_quality", "naturalness", "prompt_adherence"]:
                if random.random() > 0.5:
                    pair = {
                        "pair_id": f"{i:03d}_{criterion}",
                        "prompt": self.prompts[i],
                        "audio_A": str(self.model1_files[i]),
                        "audio_B": str(self.model2_files[i]),
                        "criterion": criterion,
                        "ground_truth": f"A={self.model1_name}, B={self.model2_name}",
                    }
                else:
                    pair = {
                        "pair_id": f"{i:03d}_{criterion}",
                        "prompt": self.prompts[i],
                        "audio_A": str(self.model2_files[i]),
                        "audio_B": str(self.model1_files[i]),
                        "criterion": criterion,
                        "ground_truth": f"A={self.model2_name}, B={self.model1_name}",
                    }
                test_pairs.append(pair)
        random.shuffle(test_pairs)
        return test_pairs

    def _create_evaluation_forms(self, output_dir, test_pairs):
        """Create evaluation forms for human listeners."""
        output_dir = Path(output_dir)
        with open(output_dir / "test_pairs_with_answers.json", "w") as f:
            json.dump(test_pairs, f, indent=2)

        listener_pairs = []
        for pair in test_pairs:
            listener_pairs.append(
                {
                    "pair_id": pair["pair_id"],
                    "prompt": pair["prompt"],
                    "audio_A": Path(pair["audio_A"]).name,
                    "audio_B": Path(pair["audio_B"]).name,
                    "criterion": pair["criterion"],
                }
            )
        with open(output_dir / "listener_evaluation_form.json", "w") as f:
            json.dump(listener_pairs, f, indent=2)

        # Optional Excel sheet
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Audio Evaluation"

            headers = [
                "Pair ID",
                "Prompt",
                "Audio A",
                "Audio B",
                "Criterion",
                "Preference (A/B)",
                "Confidence (1-5)",
                "Comments",
                "Listener ID",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            for row, pair in enumerate(listener_pairs, 2):
                ws.cell(row=row, column=1, value=pair["pair_id"])
                ws.cell(row=row, column=2, value=pair["prompt"])
                ws.cell(row=row, column=3, value=pair["audio_A"])
                ws.cell(row=row, column=4, value=pair["audio_B"])
                ws.cell(row=row, column=5, value=pair["criterion"])

            # widths
            for column in ws.columns:
                max_len = 0
                col_letter = column[0].column_letter
                for cell in column:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

            wb.save(output_dir / "audio_evaluation_form.xlsx")
        except Exception:
            print("[HumanEval] Excel form not created (openpyxl missing).")

    def _create_analysis_templates(self, output_dir):
        """Create statistical analysis templates."""
        analysis_script = r'''
import pandas as pd
import numpy as np
from scipy import stats
import matplotlib.pyplot as plt
import seaborn as sns

def analyze_human_evaluation(results_file):
    """
    Analyze human evaluation results
    Expected columns: pair_id, criterion, preference, confidence, listener_id
    """
    df = pd.read_csv(results_file)
    df['preference_numeric'] = df['preference'].map({'A': 0, 'B': 1})

    overall = {}
    for criterion in df['criterion'].unique():
        d = df[df['criterion'] == criterion]
        counts = d['preference'].value_counts()
        total = len(d)
        successes = counts.get('B', 0)  # B == model 2
        p_value = stats.binom_test(successes, total, p=0.5, alternative='two-sided')
        prop = successes / max(total, 1)
        me = 1.96 * np.sqrt(prop * (1 - prop) / max(total, 1))
        overall[criterion] = {
            'pref_A_percent': 100.0 * counts.get('A', 0) / max(total, 1),
            'pref_B_percent': 100.0 * counts.get('B', 0) / max(total, 1),
            'p_value': float(p_value),
            'significant': bool(p_value < 0.05),
            'conf_int': (float(prop - me), float(prop + me)),
            'n': int(total),
        }
    return overall

def calculate_inter_rater_reliability(df):
    """Mean pairwise Pearson correlation as a simple reliability proxy."""
    from scipy.stats import pearsonr
    pivot = df.pivot_table(index='pair_id', columns='listener_id',
                           values='preference_numeric', aggfunc='first')
    corrs = []
    cols = list(pivot.columns)
    for i in range(len(cols)):
        for j in range(i+1, len(cols)):
            valid = ~(pivot[cols[i]].isna() | pivot[cols[j]].isna())
            if valid.sum() > 0:
                c, _ = pearsonr(pivot.loc[valid, cols[i]], pivot.loc[valid, cols[j]])
                corrs.append(c)
    return float(np.mean(corrs)) if corrs else 0.0
'''
        Path(output_dir, "human_evaluation_analysis.py").write_text(analysis_script)

    def _create_experimenter_guidelines(self, output_dir):
        """Create comprehensive experimenter guidelines."""
        guidelines = '''# Human Audio Evaluation Guidelines

## Pre-Experiment Setup
- High-quality headphones/monitors, quiet room (<40 dB), calibrated volume
- 20–100+ participants depending on rigor
- Randomized, blinded A/B with confidence rating

## Protocol (summary)
1) Briefing & practice (10 min)
2) Main evaluation (30–45 min): A/B pairs, forced choice + confidence (1–5)
3) Breaks every 15–20 pairs
4) Post-eval questionnaire and debrief

## Primary Statistics
- Binomial tests (α=0.05), effect sizes (Cohen's h), CIs
- Inter-rater reliability (correlation/ICC)

## Ethics & QC
- Informed consent, right to withdraw, anonymized data
- Attention checks, response time filters, integrity checks
'''
        Path(output_dir, "human_evaluation_guidelines.md").write_text(guidelines)

    # ---------- Markdown table helper ----------
    def _metrics_table_markdown(self, statistical_results):
        import pandas as pd
        import numpy as np

        order = ["rms", "snr", "spectral_centroid", "spectral_bandwidth", "dynamic_range", "frequency_diversity"]

        rows = []
        for metric in order:
            res = statistical_results.get(metric)
            if not res:
                continue
            rows.append({
                "Metric": metric.replace("_", " ").title(),
                f"{self.model1_name} Mean ± Std": (
                    f"{res.get(f'{self.model1_name}_mean', np.nan):.4f} ± "
                    f"{res.get(f'{self.model1_name}_std',  np.nan):.4f}"
                ),
                f"{self.model2_name} Mean ± Std": (
                    f"{res.get(f'{self.model2_name}_mean', np.nan):.4f} ± "
                    f"{res.get(f'{self.model2_name}_std',  np.nan):.4f}"
                ),
                "p-value": f"{res.get('p_value', np.nan):.4g}",
                "Effect Size": res.get("effect_size", "n/a"),
            })

        df = pd.DataFrame(rows, columns=[
            "Metric",
            f"{self.model1_name} Mean ± Std",
            f"{self.model2_name} Mean ± Std",
            "p-value",
            "Effect Size",
        ])

        try:
            md_table = df.to_markdown(index=False)
        except Exception:
            headers = list(df.columns)
            line = "|" + "|".join(headers) + "|\n"
            sep   = "|" + "|".join(["---"] * len(headers)) + "|\n"
            body_lines = []
            for _, r in df.iterrows():
                body_lines.append("|" + "|".join(str(v) for v in r.values) + "|\n")
            md_table = line + sep + "".join(body_lines)

        return md_table, df

    # ----------------------------
    # Visualization & report
    # ----------------------------
    def create_comparison_plots(self, df, output_dir):
        """Create visualization plots."""
        print("[Eval] Plots …")
        output_dir = Path(output_dir)
        plt.style.use("seaborn-v0_8")
        fig, axes = plt.subplots(2, 3, figsize=(18, 12))
        fig.suptitle(f"{self.model1_name} vs {self.model2_name} Comparison", fontsize=16, fontweight="bold")

        metrics_to_plot = ["rms", "snr", "spectral_centroid", "dynamic_range", "frequency_diversity", "spectral_bandwidth"]
        for i, metric in enumerate(metrics_to_plot):
            r, c = i // 3, i % 3
            ax = axes[r, c]
            try:
                df.boxplot(column=metric, by="model", ax=ax)
                ax.set_title(metric.replace("_", " ").title())
                ax.set_xlabel("Model")
                ax.set_ylabel(metric.replace("_", " ").title())
            except Exception:
                ax.set_title(f"{metric} (unavailable)")

        plt.tight_layout()
        plt.savefig(output_dir / "model_comparison_plots.png", dpi=300, bbox_inches="tight")
        plt.close()

        # Distribution comparison
        fig, axes = plt.subplots(2, 2, figsize=(15, 10))
        fig.suptitle("Distribution Comparisons", fontsize=16, fontweight="bold")
        key_metrics = ["rms", "snr", "spectral_centroid", "dynamic_range"]
        for i, metric in enumerate(key_metrics):
            r, c = i // 2, i % 2
            ax = axes[r, c]
            try:
                d1 = df[df["model"] == self.model1_name][metric].dropna()
                d2 = df[df["model"] == self.model2_name][metric].dropna()
                ax.hist(d1, alpha=0.7, label=self.model1_name, bins=15, density=True)
                ax.hist(d2, alpha=0.7, label=self.model2_name, bins=15, density=True)
                ax.set_xlabel(metric.replace("_", " ").title())
                ax.set_ylabel("Density")
                ax.set_title(f"{metric.replace('_', ' ').title()} Distribution")
                ax.legend()
            except Exception:
                ax.set_title(f"{metric} (unavailable)")

        plt.tight_layout()
        plt.savefig(output_dir / "distribution_comparison.png", dpi=300, bbox_inches="tight")
        plt.close()

    def _build_full_html_report(self, output_dir: Path, metrics_dir: Path, viz_dir: Path,
                                stats_results: dict, diversity_df: pd.DataFrame,
                                fad_results: dict | None):
        """Create a rich report.html that embeds all metrics & visualizations."""
        # Load what we wrote to disk
        detailed_csv = metrics_dir / "detailed_metrics.csv"
        stats_csv    = metrics_dir / "statistical_comparison.csv"
        diversity_csv= metrics_dir / "diversity_metrics.csv"
        clap_sum_csv = metrics_dir / "clap_prompt_adherence_summary.csv"
        fad_json     = metrics_dir / "fad_results.json"

        # Make a quality summary (means ± std per model)
        quality_df = pd.read_csv(detailed_csv) if detailed_csv.exists() else pd.DataFrame()
        summary_rows = []
        if not quality_df.empty:
            metric_cols = ["rms","snr","spectral_centroid","spectral_bandwidth","dynamic_range","frequency_diversity"]
            for model, sub in quality_df.groupby("model"):
                row = {"Model": model}
                for m in metric_cols:
                    if m in sub.columns:
                        row[f"{m} mean"] = np.nanmean(sub[m])
                        row[f"{m} std"]  = np.nanstd(sub[m])
                summary_rows.append(row)
        summary_df = pd.DataFrame(summary_rows)

        # Other tables
        stats_df     = pd.read_csv(stats_csv) if stats_csv.exists() else pd.DataFrame()
        diversity_df = pd.read_csv(diversity_csv) if isinstance(diversity_df, pd.DataFrame) and diversity_df.empty and diversity_csv.exists() else diversity_df
        clap_df      = pd.read_csv(clap_sum_csv) if clap_sum_csv.exists() else pd.DataFrame()
        fad_data     = json.loads(fad_json.read_text()) if fad_json.exists() else fad_results

        # Render snippets
        def df_html(df: pd.DataFrame, none_msg: str):
            return df.to_html(index=False, float_format=lambda x: f"{x:.4f}") if not df.empty else f"<p>{none_msg}</p>"

        fad_html = "<p>FAD not computed.</p>"
        if fad_data:
            fad_html = f"""
<ul>
  <li><b>FAD Score:</b> {fad_data['fad_score']:.4f}</li>
  <li><b>Samples Used:</b> {fad_data['model1_samples']} vs {fad_data['model2_samples']}</li>
  <li><i>{fad_data['interpretation']}</i></li>
</ul>"""

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Audio Model Comparison Report</title>
<style>
  body {{ font-family: system-ui, -apple-system, Segoe UI, Roboto, Helvetica, Arial, sans-serif; margin: 24px; }}
  h1, h2, h3 {{ margin: 0.6em 0 0.3em; }}
  table {{ border-collapse: collapse; width: 100%; }}
  th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
  th {{ background: #f6f6f6; }}
  .mono {{ font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, "Liberation Mono", monospace; font-size:12px; }}
  img {{ max-width: 100%; height: auto; }}
  .two {{ display:grid; grid-template-columns: 1fr 1fr; gap: 16px; }}
</style>
</head>
<body>
<h1>Audio Model Comparison Report: {self.model1_name} vs {self.model2_name}</h1>
<p class="mono">Generated at: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}</p>

<h2>Fréchet Audio Distance (FAD)</h2>
{fad_html}

<h2>Quality Summary (means ± std)</h2>
{df_html(summary_df, "Quality metrics not available.")}

<h2>Statistical Comparison</h2>
{df_html(stats_df, "Statistical comparison not available.")}

<h2>Diversity Metrics</h2>
{df_html(diversity_df if isinstance(diversity_df, pd.DataFrame) else pd.DataFrame(), "Diversity metrics not available.")}

<h2>CLAP Prompt Adherence (mean by model)</h2>
{df_html(clap_df, "CLAP prompt adherence not computed.")}

<h2>Visualizations</h2>
<div class="two">
  <div><img src="visualizations/model_comparison_plots.png" alt="Boxplots"></div>
  <div><img src="visualizations/distribution_comparison.png" alt="Distributions"></div>
</div>

<hr>
<p class="mono">Artifacts in:
  <b>metrics/</b> (CSVs/JSON) and <b>visualizations/</b> (PNGs)</p>
</body>
</html>
"""
        (output_dir / "report.html").write_text(html, encoding="utf-8")
        print(f"[DONE] HTML report → {output_dir/'report.html'}")

    # ----------------------------
    # Orchestrator
    # ----------------------------
    def generate_evaluation_report(self, output_dir="evaluation_results"):
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        # categorized subfolders
        metrics_dir = output_dir / "metrics"
        viz_dir = output_dir / "visualizations"
        metrics_dir.mkdir(exist_ok=True)
        viz_dir.mkdir(exist_ok=True)

        # Objective metrics
        m1 = self.evaluate_audio_quality(self.model1_files, self.model1_name)
        m2 = self.evaluate_audio_quality(self.model2_files, self.model2_name)
        all_metrics = pd.concat([m1, m2], ignore_index=True)
        all_metrics.to_csv(metrics_dir / "detailed_metrics.csv", index=False)

        # Diversity
        d1 = self.evaluate_audio_diversity(self.model1_files, self.model1_name)
        d2 = self.evaluate_audio_diversity(self.model2_files, self.model2_name)
        diversity_df = pd.DataFrame([d1, d2])
        diversity_df.to_csv(metrics_dir / "diversity_metrics.csv", index=False)

        # FAD
        fad_results = self.evaluate_fad_score()
        if fad_results:
            with open(metrics_dir / "fad_results.json", "w") as f:
                json.dump(fad_results, f, indent=2)

        # CLAP prompt adherence (write into metrics/)
        clap_df = self.evaluate_prompt_adherence(metrics_dir)
        if clap_df is not None:
            try:
                temp = clap_df[["index", "model", "clap_similarity"]].rename(
                    columns={"clap_similarity": "clap_sim"}
                )
                temp.groupby("model")["clap_sim"].mean().to_csv(metrics_dir / "clap_prompt_adherence_summary.csv")
            except Exception:
                pass

        # Statistics
        quality_metrics = ["rms", "snr", "spectral_centroid", "spectral_bandwidth", "dynamic_range", "frequency_diversity"]
        stats_results = self.compare_models_statistically(
            all_metrics[all_metrics["model"] == self.model1_name],
            all_metrics[all_metrics["model"] == self.model2_name],
            quality_metrics,
        )
        pd.DataFrame.from_dict(stats_results, orient="index").to_csv(metrics_dir / "statistical_comparison.csv")

        # Plots → visualizations/
        self.create_comparison_plots(all_metrics, viz_dir)

        # Human evaluation kit → metrics/
        test_pairs = self.create_human_evaluation_protocol(metrics_dir)

        # Rich HTML report with everything inline
        self._build_full_html_report(output_dir, metrics_dir, viz_dir, stats_results, diversity_df, fad_results)

        return {
            "detailed_metrics": all_metrics,
            "diversity_metrics": diversity_df,
            "statistical_comparison": stats_results,
            "fad_results": fad_results,
            "human_evaluation_pairs": len(test_pairs),
        }


# ----------------------------
# CLI
# ----------------------------
def main():
    import argparse

    parser = argparse.ArgumentParser(description="Industry Standard Audio Model Evaluation")
    parser.add_argument("--model1-dir", required=True, help="Directory with AudioLDM1 samples")
    parser.add_argument("--model2-dir", required=True, help="Directory with AudioLDM2 samples")
    parser.add_argument("--prompts-file", required=True, help="File with prompts used")
    parser.add_argument("--output-dir", default="evaluation_results", help="Output directory")
    parser.add_argument("--model1-name", default="AudioLDM1", help="Name for first model")
    parser.add_argument("--model2-name", default="AudioLDM2", help="Name for second model")

    # Devices / checkpoints
    parser.add_argument("--fad-device", choices=["cpu", "gpu"], default="cpu",
                        help="Device for FAD (TF VGGish). 'cpu' avoids CUDA/cuDNN issues")
    parser.add_argument("--clap-checkpoint", default="laion/clap-htsat-unfused",
                        help="HF CLAP checkpoint (e.g., laion/clap-htsat-unfused or laion/clap-htsat-fused)")
    parser.add_argument("--clap-device", default="cuda", help="Device for CLAP (cuda or cpu)")

    args = parser.parse_args()

    evaluator = AudioModelEvaluator(
        args.model1_dir,
        args.model2_dir,
        args.prompts_file,
        args.model1_name,
        args.model2_name,
        fad_device=args.fad_device,
        clap_checkpoint=args.clap_checkpoint,
        clap_device=args.clap_device,
    )

    results = evaluator.generate_evaluation_report(args.output_dir)

    print("\nEvaluation Summary:")
    print(f"Processed {len(results['detailed_metrics'])} rows of metrics")
    print(f"Results saved to: {args.output_dir}")


if __name__ == "__main__":
    main()